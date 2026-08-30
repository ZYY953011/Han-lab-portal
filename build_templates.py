# -*- coding: utf-8 -*-
"""
生成课题组信息收集模板（10 张独立 Excel 表）。
字段与 /workspace/assets/js/data/*.js 中网站数据结构一致，
组员填完后交回，即可由管理员替换为网站数据。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

OUT = "/workspace/信息收集模板"
os.makedirs(OUT, exist_ok=True)

# ---------- 通用样式 ----------
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")      # 科研蓝
HEAD_FONT  = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
HEAD_FILL  = PatternFill("solid", fgColor="DDEBF7")
EX_FILL    = PatternFill("solid", fgColor="FFF2CC")      # 示例行 浅黄
NOTE_FILL  = PatternFill("solid", fgColor="E2EFDA")
WRAP       = Alignment(wrap_text=True, vertical="top")
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build(filename, title, instructions, headers, example, validations=None, blank_rows=15):
    """headers: list of (中文列名, 填写说明/留空)
       example: list of 示例值（与 headers 等长）
       validations: list of (列索引0基, ["选项1","选项2",...])"""
    wb = Workbook()
    ws = wb.active
    ws.title = "填报表"

    ncol = len(headers)
    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # 说明行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(row=2, column=1, value="填写说明：" + instructions)
    c.font = Font(name="微软雅黑", size=9, color="375623")
    c.fill = NOTE_FILL; c.alignment = WRAP
    ws.row_dimensions[2].height = 30

    # 表头行（第3行）
    for j, (hname, hnote) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=j, value=hname)
        cell.font = HEAD_FONT; cell.fill = HEAD_FILL
        cell.alignment = CENTER; cell.border = BORDER
        if hnote:
            cell.comment = Comment(hnote, "模板")
    ws.row_dimensions[3].height = 30

    # 示例行（第4行，黄色，交回前可删除）
    if example:
        for j, val in enumerate(example, start=1):
            cell = ws.cell(row=4, column=j, value=val)
            cell.fill = EX_FILL; cell.alignment = WRAP; cell.border = BORDER
            cell.font = Font(name="微软雅黑", size=9, color="7F6000")
        ws.row_dimensions[4].height = 42

    # 空白填写行
    start = 5 if example else 4
    for r in range(start, start + blank_rows):
        for j in range(1, ncol + 1):
            cell = ws.cell(row=r, column=j)
            cell.alignment = WRAP; cell.border = BORDER

    # 下拉校验
    if validations:
        for col_idx, opts in validations:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(opts) + '"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            col_letter = ws.cell(row=4 if example else 3, column=col_idx + 1).column_letter
            dv.add(f"{col_letter}{start}:{col_letter}{start+blank_rows-1}")

    # 列宽
    for j in range(1, ncol + 1):
        maxlen = len(str(headers[j-1][0]))
        for r in range(4 if example else 3, start + blank_rows):
            v = ws.cell(row=r, column=j).value
            if v:
                maxlen = max(maxlen, min(len(str(v)), 20))
        ws.column_dimensions[ws.cell(row=3, column=j).column_letter].width = min(max(maxlen + 4, 12), 40)

    ws.freeze_panes = "A4" if example else "A3"
    path = os.path.join(OUT, filename)
    wb.save(path)
    print("已生成:", filename)


# ============ 1. 项目 ============
build(
    "01-项目信息表.xlsx",
    "项目信息表（对应网站「项目管理」模块）",
    "每个项目一行；编号请唯一（P1、P2…）。关联字段填对应模块的编号，多个用逗号分隔。黄色行为示例，正式提交前删除。",
    headers=[
        ("项目编号", "唯一，字母+数字，如 P1"),
        ("项目名称", ""),
        ("项目负责人", ""),
        ("参与成员", "多人用逗号分隔，如 张三,李四"),
        ("项目来源", "如 国家自然科学基金面上项目"),
        ("项目外部编号", "如 NSFC-32171800"),
        ("起始时间", "年-月，如 2022-01"),
        ("结束时间", "年-月，如 2025-12"),
        ("经费", "如 60 万元"),
        ("当前状态", "见下拉"),
        ("当前阶段", "如 基因功能验证"),
        ("进度百分比", "数字 0-100"),
        ("项目目标", ""),
        ("OKR", "每条一行，可多行"),
        ("本月进展", ""),
        ("下一步计划", ""),
        ("风险和问题", ""),
        ("关联实验方法ID", "如 M1,M2"),
        ("关联样品ID", "如 S1,S2"),
        ("关联数据ID", "如 D1"),
        ("关联报账ID", "如 E1"),
        ("关联成果ID", "如 A1"),
    ],
    example=["P1", "小麦抗旱基因挖掘与分子机制研究", "张明远 教授", "张明远,李文,赵磊,陈思",
             "国家自然科学基金面上项目", "NSFC-32171800", "2022-01", "2025-12", "60 万元",
             "进行中", "基因功能验证", 72, "挖掘 2–3 个抗旱主效基因", "O：完成抗旱资源筛选 KR：鉴定50份抗旱种质",
             "本月完成转录组第3批取样", "下月开展VIGS验证", "转基因材料构建周期长", "M1,M2,M5",
             "S1,S2,S3", "D1,D2", "E1,E2", "A1"],
    validations=[(9, ["准备中", "进行中", "暂停", "结题中", "已结题"])],
)

# ============ 2. 实验方法/SOP ============
build(
    "02-实验方法表.xlsx",
    "实验方法 / SOP 表（对应网站「实验方法」模块）",
    "每个方法一行；编号唯一（M1、M2…）。多值字段在单元格内换行或逗号分隔。黄色行为示例。",
    headers=[
        ("方法编号", "唯一，如 M1"),
        ("方法名称", ""),
        ("所属类型", "如 生理指标 / 分子生物学"),
        ("适用实验", ""),
        ("适用项目", "如 P1,P3"),
        ("编写人", ""),
        ("当前版本", "如 v1.0"),
        ("最近更新时间", "年-月-日"),
        ("实验目的", ""),
        ("所需材料", "逗号分隔"),
        ("所需仪器", "逗号分隔"),
        ("操作步骤", "每条一行"),
        ("关键参数", "逗号分隔"),
        ("注意事项", "每条一行"),
        ("常见错误", "每条一行"),
        ("失败案例", "每条一行"),
        ("推荐经验", "每条一行"),
        ("参考文献", "每条一行"),
        ("附件链接", "格式：名称|URL"),
        ("视频链接", ""),
        ("版本历史", "每行一条：版本|日期|修改人|说明"),
    ],
    example=["M1", "小麦叶片光合速率测定（LI-6400）", "生理指标", "逆境生理、光合特性评价", "P1,P3",
             "陈思", "v2.1", "2024-03-12", "测定小麦叶片净光合速率等参数",
             "小麦旗叶,蒸馏水,夹叶器", "LI-6400 便携式光合仪", "开机预热30分钟\n选取健康旗叶\n记录3次取平均",
             "光强1200,温度25℃,CO₂400ppm", "测定前叶片须适应光强10分钟", "叶室遮光不均导致Pn偏低",
             "2023-07一批数据异常已作废", "固定同一操作人员", "Long S P,2009", "光合测定记录表|https://docs.example.com/pn_record",
             "https://video.example.com/li6400_demo", "v2.1|2024-03-12|陈思|补充高温天注意事项"],
)

# ============ 3. 样品 ============
build(
    "03-样品信息表.xlsx",
    "样品信息表（对应网站「样品管理」模块）",
    "每个样品一行；编号唯一（S1、S2…）。位置字段请填到最小单元，便于查找。黄色行为示例。",
    headers=[
        ("样品ID", "唯一，如 S1"),
        ("样品编码", "外部编码，如 WH-2024-001"),
        ("样品名称", ""),
        ("样品类型", "如 种子/植物组织/土壤/核酸"),
        ("来源", ""),
        ("所属项目", "如 P1"),
        ("负责人", ""),
        ("制备/采样日期", "年-月-日"),
        ("保存条件", "如 -80℃冻存 / 4℃干燥"),
        ("所在楼栋", ""),
        ("房间", ""),
        ("冰箱/柜子", ""),
        ("层", ""),
        ("盒号", ""),
        ("具体位置", "如 超低温冰箱2号/中层/盒12"),
        ("总数量", "数字"),
        ("剩余数量", "数字"),
        ("单位", "如 粒/管/袋"),
        ("当前状态", "见下拉"),
        ("最近使用人", ""),
        ("最近使用时间", "年-月-日"),
        ("备注", ""),
        ("二维码编号", "如 QR-S1（留空也可）"),
    ],
    example=["S1", "WH-2024-001", "小麦抗旱种质种子", "种子", "种质资源库引种", "P1", "陈思",
             "2024-01-10", "4℃ 干燥", "农科楼", "B205 种子室", "种子柜 A", "第2层", "盒07",
             "种子柜A/第2层/盒07", 500, 120, "粒", "在库", "陈思", "2024-03-08",
             "含50份抗旱与对照种质", "QR-S1"],
    validations=[(18, ["在库", "使用中", "已耗尽", "已转移", "已废弃"])],
)

# ============ 4. 数据索引 ============
build(
    "04-数据索引表.xlsx",
    "数据索引表（对应网站「数据管理」模块，只存链接不存大文件）",
    "每条数据一行；编号唯一（D1、D2…）。原始大文件放在服务器/网盘，这里只填链接。黄色行为示例。",
    headers=[
        ("数据ID", "唯一，如 D1"),
        ("数据名称", ""),
        ("所属项目", "如 P1"),
        ("实验名称", ""),
        ("实验日期", "年-月-日"),
        ("材料", ""),
        ("处理", ""),
        ("重复", "如 3生物学重复"),
        ("测量变量", ""),
        ("操作者", ""),
        ("仪器", ""),
        ("原始数据链接", "NAS/网盘/对象存储URL"),
        ("清洗数据链接", ""),
        ("分析代码链接", "如 GitHub 地址"),
        ("图表链接", ""),
        ("结果说明", ""),
        ("文件版本", "如 v1.0"),
        ("更新时间", "年-月-日"),
        ("存储位置", "如 nas://group/xm"),
        ("外部ID", "如 NAS-XM-001"),
    ],
    example=["D1", "小麦抗旱转录组原始测序数据", "P1", "抗旱转录组", "2024-03-05", "小麦旗叶",
             "PEG模拟干旱0/6/12/24h", "3生物学重复", "基因表达量FPKM", "赵磊", "Illumina NovaSeq",
             "https://nas.example.edu.cn/share/xm/rnaseq_raw", "https://nas.example.edu.cn/share/xm/rnaseq_clean",
             "https://github.com/group/xm_rnaseq", "https://nas.example.edu.cn/share/xm/figures/pca.png",
             "原始fastq约120GB存于学校NAS", "v1.0", "2024-03-15", "nas://group/xm/rnaseq", "NAS-XM-001"],
)

# ============ 5. 报账 ============
build(
    "05-报账记录表.xlsx",
    "报账记录表（对应网站「报账与采购」模块）",
    "每笔报销一行；编号唯一（E1、E2…）。金额填数字。黄色行为示例。",
    headers=[
        ("报账编号", "唯一，如 E1"),
        ("日期", "年-月-日"),
        ("报账人", ""),
        ("所属项目", "如 P1"),
        ("费用类别", "见下拉"),
        ("金额", "数字，单位元"),
        ("用途", ""),
        ("发票状态", "见下拉"),
        ("材料是否齐全", "如 齐全 / 缺失合同"),
        ("当前状态", "见下拉"),
        ("提交日期", "年-月-日"),
        ("报销完成日期", "年-月-日，未完成留空"),
        ("备注", ""),
    ],
    example=["E1", "2024-03-02", "赵磊", "P1", "材料采购", 8600, "Trizol、引物、试剂盒等耗材",
             "已开票", "齐全", "已报销", "2024-03-05", "2024-03-20", "国库集中支付"],
    validations=[
        (4, ["差旅报销", "材料采购", "劳务费", "会议费", "打车费", "维修费", "场地费"]),
        (7, ["已开票", "待开票", "不需发票"]),
        (9, ["待整理", "待提交", "已提交", "退回修改", "审核中", "已报销"]),
    ],
)

# ============ 6. 组会 ============
build(
    "06-组会记录表.xlsx",
    "组会记录表（对应网站「组会安排与记录」模块）",
    "每次组会一行；编号用日期（MT-2024-03-22）。行动事项最多列3条，不够可加行。黄色行为示例。",
    headers=[
        ("组会编号", "如 MT-2024-03-22"),
        ("组会日期", "年-月-日"),
        ("地点", ""),
        ("主持人", ""),
        ("汇报人", ""),
        ("汇报主题", ""),
        ("所属项目", "如 P1"),
        ("PPT链接", ""),
        ("文档链接", ""),
        ("主要进展", ""),
        ("当前问题", ""),
        ("讨论内容", ""),
        ("老师建议", ""),
        ("行动事项1-任务", ""),
        ("行动事项1-负责人", ""),
        ("行动事项1-截止", "年-月-日"),
        ("行动事项1-检查", "年-月-日"),
        ("行动事项2-任务", ""),
        ("行动事项2-负责人", ""),
        ("行动事项2-截止", "年-月-日"),
        ("行动事项2-检查", "年-月-日"),
        ("行动事项3-任务", ""),
        ("行动事项3-负责人", ""),
        ("行动事项3-截止", "年-月-日"),
        ("行动事项3-检查", "年-月-日"),
    ],
    example=["MT-2024-03-22", "2024-03-22", "农科楼B308", "张明远", "陈思",
             "小麦抗旱转录组进展与候选基因汇报", "P1", "https://docs.example.com/mt0322_ppt",
             "https://docs.example.com/mt0322_note", "完成转录组分析，筛选12个候选基因",
             "转基因材料构建周期长", "讨论优先验证Top3", "老师建议先补VIGS验证",
             "完成Top3基因VIGS载体构建", "赵磊", "2024-04-10", "2024-04-12",
             "补充干旱表型拍照存档", "陈思", "2024-04-05", "2024-04-08", "", "", "", ""],
)

# ============ 7. 学生培养（扁平：每生8阶段行）============
build(
    "07-学生培养表.xlsx",
    "学生培养表（对应网站「学生培养」模块）",
    "每位学生占8行（入组→毕业8个阶段各一行）。前两列（学生ID/姓名等）每行都填相同。黄色行为示例。",
    headers=[
        ("学生ID", "唯一，如 ST1"),
        ("姓名", ""),
        ("学号", ""),
        ("培养类型", "见下拉"),
        ("入学时间", "年-月"),
        ("入组时间", "年-月"),
        ("导师", ""),
        ("研究方向", ""),
        ("所属项目", "如 P1"),
        ("预计毕业", "年-月"),
        ("阶段名称", "见下拉，8个阶段各一行"),
        ("计划日期", "年-月"),
        ("实际日期", "年-月，未完成留空"),
        ("阶段状态", "见下拉"),
        ("所需材料", ""),
        ("材料链接", ""),
        ("导师意见", ""),
        ("是否完成", "是/否"),
        ("备注", ""),
    ],
    example=[
        "ST1","陈思","2021110023","博士生","2021-09","2021-10","张明远","小麦抗旱基因功能解析","P1","2025-06",
        "入组","2021-10","2021-10","已完成","入组登记表","","","是","",
    ],
    validations=[
        (3, ["博士生", "硕士生", "本科生", "博士后"]),
        (10, ["入组", "培养计划", "开题", "中期考核", "预答辩", "毕业答辩", "学位材料", "毕业"]),
        (13, ["未开始", "准备中", "已完成", "需要修改"]),
        (17, ["是", "否"]),
    ],
    blank_rows=40,
)

# ============ 8. 成员 ============
build(
    "08-成员信息表.xlsx",
    "成员信息表（对应网站「成员」模块）",
    "每位成员一行；编号唯一（MB1、MB2…）。联系方式为公开信息，敏感私人信息勿填。黄色行为示例。",
    headers=[
        ("成员ID", "唯一，如 MB1"),
        ("姓名", ""),
        ("身份", "见下拉"),
        ("入组时间", "年-月"),
        ("预计毕业时间", "年-月，在职教师留空"),
        ("研究方向", ""),
        ("当前项目", "如 P1,P2"),
        ("擅长技能", "逗号分隔"),
        ("可提供帮助", ""),
        ("联系方式", "课题组邮箱/工位电话，勿填私人手机"),
        ("工位", ""),
        ("当前状态", "见下拉"),
        ("毕业去向", "已毕业成员填，其余留空"),
    ],
    example=["MB1", "张明远", "教师", "2015-03", "", "作物生理生态与产量形成", "P1,P2",
             "课题设计,栽培生理,论文指导", "可指导实验设计与投稿", "zhangmy@univ.edu.cn",
             "农科楼B402", "在职", ""],
    validations=[
        (2, ["教师", "博士后", "博士生", "硕士生", "本科生", "已毕业成员"]),
        (11, ["在职", "在读", "已毕业"]),
    ],
)

# ============ 9. 研究成果 ============
build(
    "09-研究成果表.xlsx",
    "研究成果表（对应网站「成果」模块）",
    "每项成果一行；编号唯一（A1、A2…）。影响因子填数字，无则填0。黄色行为示例。",
    headers=[
        ("成果ID", "唯一，如 A1"),
        ("成果类型", "见下拉"),
        ("标题", ""),
        ("作者", "逗号分隔"),
        ("通讯作者", ""),
        ("期刊/载体", ""),
        ("年份", "数字"),
        ("DOI", ""),
        ("分区/类别", "如一区/中文核心/专利/软著"),
        ("影响因子", "数字，无则0"),
        ("所属项目", "如 P1"),
        ("主要学生", "逗号分隔"),
        ("论文链接", ""),
        ("PDF链接", ""),
    ],
    example=["A1", "论文", "TaNAC67 调控小麦抗旱性的分子机制", "陈思,赵磊,张明远", "张明远",
             "Journal of Experimental Botany", 2024, "10.1093/jxb/abcd123", "一区", 6.9,
             "P1", "陈思", "https://doi.org/10.1093/jxb/abcd123", "https://docs.example.com/pdf/a1"],
    validations=[(1, ["论文", "专利", "软件著作权", "获奖", "学术会议", "基金", "品种", "标准", "数据集", "学生竞赛"])],
)

# ============ 10. 学习资源 ============
build(
    "10-学习资源表.xlsx",
    "学习资源表（对应网站「学习资源」模块）",
    "每条资源一行；编号唯一（R1、R2…）。推荐程度用★数量表示。黄色行为示例。",
    headers=[
        ("资源ID", "唯一，如 R1"),
        ("标题", ""),
        ("分类", ""),
        ("简介", ""),
        ("适合谁看", ""),
        ("建议什么时候看", ""),
        ("预计阅读时间(分钟)", "数字"),
        ("链接", ""),
        ("推荐程度", "如 ★★★★★"),
    ],
    example=["R1", "新生入组手册（必读）", "新生入门", "课题组规矩、安全、常用账号与日常流程。",
             "新入组研究生/本科生", "入组第一周", 20, "https://docs.example.com/res/freshman", "★★★★★"],
)

print("\n全部模板已生成至:", OUT)
