"""生成飞书多维表格建表模板（20 列：含自购备案流程、报销方式、发票编号等）
对照实际流程：差旅报销 / 三助报销 / 自购实验材料费 / 工人报销 / 校内转账 / 校外教师劳务费
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ===== Sheet 1: 字段定义 =====
ws1 = wb.active
ws1.title = "字段定义"

# 字段定义：(列名, 类型, 必填, 选项, 说明)
fields = [
    ("报销编号",         "自动编号",     "是", "—",                                        "飞书自动生成"),
    ("报销人",           "单选",         "是", "李佳翔/任浩/张帆/陈瑞雪/…(全部组员)",         "可后续升级为成员表关联记录"),
    ("花费日期",         "日期",         "是", "格式 YYYY-MM-DD",                            "实际消费发生日期"),
    ("费用大类",         "单选",         "是",
                       "差旅报销 / 三助报销 / 自购实验材料费 / 工人报销 / 校内转账 / 校外教师劳务费",
                       "决定后续审核路径——尤其自购实验材料费需多走学院备案"),
    ("报销方式",         "单选",         "是", "个人垫付 / 对公转账",                          "与网站 expenseFlows 对应"),
    ("有无发票",         "单选",         "是", "有 / 无 / 部分",                              "决定走三助报销还是有发票流程"),
    ("发票或截图编号",   "多行文本",     "否", "示例：3、4、5 或 15、16",                       "对应文件夹里凭证图片的编号，方便核账"),
    ("付款凭证",         "附件",         "否", "PDF/JPG/PNG",                                  "发票、转账截图、订单截图；编号与上面列对应"),
    ("金额（元）",       "货币",         "是", "单位:元,保留 2 位小数",                        "—"),
    ("项目标注",         "单选",         "是", "黄淮海小麦项目 / 单产项目 / 联合基金项目 / …",   "导师汇总表里的项目列"),
    ("详细内容",         "多行文本",     "是", "—",
                       "材料类:写了名称与数量；劳务类:写工人/教师姓名与工时或工作量"),
    ("备注",             "多行文本",     "否", "—",                                            "工人姓名、特殊情况说明"),
    ("自购备案状态",     "单选",         "否", "不需要 / 待申请 / 已申请 / 已拿到备案",
                       "仅费用大类=自购实验材料费 时才用；其他类别选「不需要」"),
    ("自购备案编号",     "单行文本",     "否", "—",                                            "拿到学院备案证明后填写"),
    ("审核状态",         "单选",         "是", "待审核 / 审核中 / 已通过 / 退回修改",            "由审核人(你)修改"),
    ("审核意见",         "多行文本",     "否", "—",                                            "退回时必填,说明原因"),
    ("下单状态",         "单选",         "是", "待下单 / 已下单",                              "由下单人修改；自购实验材料需拿到备案后才能标「已下单」"),
    ("导师审批",         "单选",         "是", "待审批 / 已批 / 驳回",                          "由项目负责人(导师)修改"),
    ("提交时间",         "创建时间",     "自动", "—",                                          "飞书自动记录"),
    ("周期",             "单选",         "是", "2026.6–2026.8 / 2026.9–2026.11 / …",            "每轮报账一个周期,用于按月统计"),
]

# 样式
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
type_fill   = PatternFill("solid", fgColor="EAF2FB")
key_fill    = PatternFill("solid", fgColor="FFF1E0")  # 高亮关键字段
hint_fill   = PatternFill("solid", fgColor="FFF8E1")
thin = Side(border_style="thin", color="C0C0C0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["列号", "列名（飞书里这样填）", "字段类型", "必填", "选项 / 格式", "说明"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.fill = header_fill; c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# 关键字段高亮（费用大类、自购备案状态、发票编号、报销方式）
key_columns = {4, 7, 13}

for i, (name, typ, req, opt, desc) in enumerate(fields, 1):
    row = i + 1
    ws1.cell(row=row, column=1, value=i)
    ws1.cell(row=row, column=2, value=name)
    type_cell = ws1.cell(row=row, column=3, value=typ)
    type_cell.fill = type_fill
    ws1.cell(row=row, column=4, value=req)
    opt_cell = ws1.cell(row=row, column=5, value=opt)
    desc_cell = ws1.cell(row=row, column=6, value=desc)
    desc_cell.fill = hint_fill
    if 4 in key_columns:
        opt_cell.fill = key_fill
    for col in range(1, 7):
        cell = ws1.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 列宽
widths = [6, 22, 12, 8, 50, 40]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height = 26

# ===== Sheet 2: 示例数据（覆盖 6 类场景） =====
ws2 = wb.create_sheet("示例数据（覆盖6类）")

sample_headers = [f[0] for f in fields]
samples = [
    # 1. 差旅报销
    ["—", "李瑞",      "2026-07-12", "差旅报销",         "个人垫付", "有",     "13",  "(传PDF)", "90.00",
     "黄淮海小麦项目", "李瑞一人去长武车费",                       "—",      "不需要", "",        "已通过", "",        "已下单", "已批", "2026-07-13 09:00", "2026.6–2026.8"],
    # 2. 三助报销
    ["—", "燕春",      "2026-06-29", "三助报销",         "个人垫付", "无",     "3、4、5", "(传图片)", "101.00",
     "黄淮海小麦项目", "买塑料袋、装钢卷尺、钻土、水",                "—",      "不需要", "",        "已通过", "",        "已下单", "已批", "2026-06-30 14:20", "2026.6–2026.8"],
    # 3. 自购实验材料费（走学院备案）
    ["—", "赵磊",      "2026-07-15", "自购实验材料费",   "对公转账", "有",     "6",      "(传PDF)", "197.90",
     "黄淮海小麦项目", "网袋一批",                                   "—",      "已拿到备案", "XY-2026-021", "审核中", "待补验货证明", "待下单", "待审批", "2026-07-16 10:15", "2026.6–2026.8"],
    # 4. 工人报销
    ["—", "燕春",      "2026-07-19", "工人报销",         "对公转账", "无",     "—",      "(传劳务发放表)", "114.00",
     "黄淮海小麦项目", "茎秆-任浩的地(男工15一小时*6+1小时*1人+女工12一小时*1人+2小时)", "苟阿姨", "不需要", "",        "已通过", "",        "已下单", "已批", "2026-07-20 11:30", "2026.6–2026.8"],
    # 5. 校内转账
    ["—", "刘钺",      "2026-07-16", "校内转账",         "对公转账", "有",     "28",     "(传校内发票)", "16.00",
     "联合基金项目",   "买水",                                       "—",      "不需要", "",        "已通过", "",        "已下单", "已批", "2026-07-16 16:40", "2026.6–2026.8"],
    # 6. 校外教师劳务费
    ["—", "陈瑞雪",    "2026-08-05", "校外教师劳务费",   "对公转账", "有",     "—",      "(传劳务表+通知)", "2000.00",
     "联合基金项目",   "外审专家答辩评审",                            "需提前报批", "不需要", "",   "已通过", "",        "已下单", "已批", "2026-08-06 09:00", "2026.6–2026.8"],
]

# 表头
for col, h in enumerate(sample_headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.fill = header_fill; c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# 数据
for ridx, row_data in enumerate(samples, 2):
    for cidx, v in enumerate(row_data, 1):
        cell = ws2.cell(row=ridx, column=cidx, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if cidx == 4:  # 费用大类列高亮
            cell.fill = key_fill

for i in range(1, len(sample_headers) + 1):
    ws2.column_dimensions[get_column_letter(i)].width = 16
ws2.row_dimensions[1].height = 32
for r in range(2, 8):
    ws2.row_dimensions[r].height = 60

# ===== Sheet 3: 6 类流程对照（飞书里"自购备案"流程最复杂） =====
ws3 = wb.create_sheet("6类流程对照")
flow_data = [
    ["费用大类",         "报销方式",  "有无发票",  "是否走学院备案",  "下一步关键动作",                                              "对应飞书单选项"],
    ["差旅报销",         "个人垫付",  "有",         "否",            "审核→下单→导师审批",                                            "差旅报销"],
    ["三助报销",         "个人垫付",  "无",         "否",            "审核→下单(走三助发放)→导师审批",                                  "三助报销"],
    ["自购实验材料费",   "对公转账",  "有",         "✅ 必须",        "审核→下单人申请自购备案→拿到备案后下报销单→财务室→导师",         "自购实验材料费"],
    ["工人报销",         "对公转账",  "无",         "否",            "审核→下单(劳务发放表+对公转账)→导师审批",                         "工人报销"],
    ["校内转账",         "对公转账",  "有",         "否",            "审核→下单(对公转账)→导师审批",                                    "校内转账"],
    ["校外教师劳务费",   "对公转账",  "有",         "否",            "审核→下单(代扣个税)→导师审批",                                    "校外教师劳务费"],
]
for ridx, row_data in enumerate(flow_data, 1):
    for cidx, v in enumerate(row_data, 1):
        cell = ws3.cell(row=ridx, column=cidx, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        if ridx == 1:
            cell.fill = header_fill
            cell.font = header_font
        elif ridx == 4:  # 自购实验材料费行高亮
            cell.fill = key_fill
            cell.font = Font(bold=True)

widths3 = [22, 14, 12, 16, 38, 22]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 26

out = "/workspace/飞书表格模板-20列.xlsx"
wb.save(out)
print(f"已生成: {out}")
print(f"Sheet 数: {len(wb.sheetnames)} -> {wb.sheetnames}")